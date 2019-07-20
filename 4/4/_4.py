import re ,math,openpyxl,os
import tweepy,platform,time 
from tweepy import OAuthHandler 
from textblob import TextBlob
from openpyxl import Workbook

  
class TwitterClient(object): 
    ''' 
    Generic Twitter Class for sentiment analysis. 
    '''
    def Create_Dir(self, dir_name):
            if not os.path.exists("data"):
                try:
                    os.mkdir("data")
                    print("Created directory 'data'")
                except:
                    print("Unable to create directory 'data': Directory already exists")
            else:
                print("Unable to create directory 'data': Directory already exists")

            if not os.path.exists("data/data_" + dir_name):
                try:
                    os.mkdir("data/data_" + dir_name)
                    print("Created directory 'data/data_" + dir_name + "'")
                except:
                    print("Unable to create directory 'data/data_" + dir_name + "': Directory already exists")
            else:
                print("Unable to create directory 'data/data_" + dir_name + "': Directory already exists")   
             # Adding path.
            if not os.getcwd() in os.get_exec_path():
            # print('adding path')
                if platform.system() == "Windows":
                    os.environ["PATH"] = os.environ["PATH"] + ";" + os.getcwd()
                else:
                    os.environ["PATH"] = os.environ["PATH"] + ":" + os.getcwd()

            
      
    def __init__(self): 
        ''' 
        Class constructor or initialization method. 
        '''
        # keys and tokens from the Twitter Dev Console 
        consumer_key = 'Y2x9rnSVirom9p4aP4j4GmWGy'
        consumer_secret = 'peCyxeqB68YvMrcdjXilHbzbWe3KXZKD4cE6NzHnjxTd2GCm9u'
        access_token = '2186053585-zX6VlzWtTr9nNg72SXk9q0TWe6yV6VDyI0TCaxF'
        access_token_secret =  'T3NAV6vXeOzHXLwBRulUXyBxRQUP8cjdbepFkeFzQyMgh'
  
        # attempt authentication 
        try: 
            # create OAuthHandler object 
            self.auth = OAuthHandler(consumer_key, consumer_secret) 
            # set access token and secret 
            self.auth.set_access_token(access_token, access_token_secret) 
            # create tweepy API object to fetch tweets 
            self.api = tweepy.API(self.auth) 
        except: 
            print("Error: Authentication Failed") 
  
    def clean_tweet(self, tweet): 
        ''' 
        Utility function to clean tweet text by removing links, special characters 
        using simple regex statements. 
        '''
        return ' '.join(re.sub("(@[A-Za-z0-9]+)|([^0-9A-Za-z \t])|(\w+:\/\/\S+)", " ", tweet).split()) 
  
    def get_tweet_sentiment(self, tweet): 
        ''' 
        Utility function to classify sentiment of passed tweet 
        using textblob's sentiment method 
        '''
        # create TextBlob object of passed tweet text 
        analysis = TextBlob(self.clean_tweet(tweet)) 
        # set sentiment 
        if analysis.sentiment.polarity > 0: 
            return 'positive'
        elif analysis.sentiment.polarity == 0: 
            return 'neutral'
        else: 
            return 'negative'
  
    def get_tweets(self, query, count = 1000,lang='en'): 
        ''' 
        Main function to fetch tweets and parse them. 
        '''
        # empty list to store parsed tweets 
        
        tweets = []   
        try: 
            # call twitter api to fetch tweets 
            fetched_tweets = self.api.search(q = query, count = count)   
            # parsing tweets one by one 
            for tweet in fetched_tweets: 
                # empty dictionary to store required params of a tweet 
                parsed_tweet = {} 
  
                # saving text of tweet 
                parsed_tweet['text'] = tweet.text 
                # saving sentiment of tweet 
                parsed_tweet['sentiment'] = self.get_tweet_sentiment(tweet.text) 
  
                # appending parsed tweet to tweets list 
                if tweet.retweet_count > 0: 
                    # if tweet has retweets, ensure that it is appended only once 
                    if parsed_tweet not in tweets: 
                        tweets.append(parsed_tweet) 
                else: 
                    tweets.append(parsed_tweet) 
  
            # return parsed tweets 
            return tweets
        except tweepy.TweepError as e: 
            # print error (if any) 
            print("Error : " + str(e))        
  
def main(): 
    # creating object of TwitterClient Class 
            api = TwitterClient() 
    # calling function to get tweets 
            query = str(input("Enter keyword to search for: "))

            tweets = api.get_tweets(query, count = 1000, lang = 'en') 

            api.Create_Dir(query)
            time.sleep(5)
            print("\n\nStarting Scrapping Twitter")        
            
            file_path = "data/data_" + query       
           # Create a workbook for excel
            tag_File = file_path + "/" + query + "_Twitter.xlsx" 
           
            wb = openpyxl.Workbook()
            ws_Pos = wb.create_sheet(title="Positive")
            col = 'A'
            row = 1

    # picking positive tweets from tweets 
            ptweets = [tweet for tweet in tweets if tweet['sentiment'] == 'positive'] 
    # percentage of positive tweets 
            print("Positive tweets percentage: {} %".format(100*len(ptweets)/len(tweets))) 
            time.sleep(5)
    # picking negative tweets from tweets 
            ntweets = [tweet for tweet in tweets if tweet['sentiment'] == 'negative'] 
    # percentage of negative tweets 
            print("Negative tweets percentage: {} %".format(100*len(ntweets)/len(tweets))) 
            time.sleep(5)
    # percentage of neutral tweets 
            nuatral = (len(tweets) - len(ntweets))- len(ptweets)
            print("Neutral tweets percentage: {} % ".format(100*(len(tweets) - len(ntweets)- len(ptweets))/len(tweets))) 
            time.sleep(5)
    # printing  positive tweets 
            print("\n\nPositive tweets:") 
            for tweet in ptweets: 
                print(tweet['text']) 
                ws_Pos['A' + str(row)] = tweet['text']
                row += 1 
            time.sleep(5)
    # printing  negative tweets 
            ws_Neg = wb.create_sheet(title="Negative")
            col = 'A'
            row = 1
            print("\n\nNegative tweets:") 
            for tweet in ntweets: 
                print(tweet['text']) 
                ws_Neg['A' + str(row)] = tweet['text']
                row += 1 
            time.sleep(5)

            wb.save(tag_File)
            print("\nClose twitter!!!")
if __name__ == "__main__": 
    # calling main function  
    main() 
    print("Stopping Scrapper...")